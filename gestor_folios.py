import os
import re
import sys
import threading
from datetime import datetime, date, timedelta

import fitz  # PyMuPDF
import io
import csv
import pandas as pd
import pyperclip
import pythoncom            # Para inicializar COM en hilos secundarios
import win32com.client
from bs4 import BeautifulSoup
from google.cloud import vision
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTextEdit, QLabel,
    QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QGridLayout,
    QFrame, QProgressBar, QGraphicsDropShadowEffect
)
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtGui import QFont, QColor

# ─── CARGAR CONFIGURACIÓN ───────────────────────────────────────
load_dotenv()

# Rutas base
RUTA_BASE_PDFS = os.getenv("RUTA_BASE_PDFS", r"path/to/Acuses")
RUTA_EXCELS    = os.getenv("RUTA_EXCELS",    r"path/to/Excels")

# Remitente para extracción de precios
SENDER_NAME    = os.getenv("SENDER_NAME",  "nombre")
SENDER_EMAIL   = os.getenv("SENDER_EMAIL", "email")

# Google Cloud Vision configura via GOOGLE_APPLICATION_CREDENTIALS env var

# ─── ESTILOS GUI ────────────────────────────────────────────────
COLOR_FONDO_VENTANA = "#f4f4f4"
COLOR_LOG_BG        = "#1E1E1E"
COLOR_LOG_FG        = "#D4D4D4"
FUENTE_LOG          = QFont("Consolas", 11)
FUENTE_BASE         = QFont("Segoe UI", 11)
COLOR_BOTON         = "#007bff"
COLOR_BOTON_HOVER   = "#005bbf"
COLOR_ERROR         = "#d32f2f"

# ─── UTILIDADES DE CONSOLA ─────────────────────────────────────
_BAR_FULL   = "■"
_BAR_EMPTY  = "·"
_BAR_LEN    = 30
_SEP        = "━" * 70
script_start_time = datetime.now().timestamp()
station_timer     = None

def _fmt_dur(sec: float) -> str:
    m, s = divmod(int(sec), 60)
    h, m = divmod(m, 60)
    if h: return f"{h} h {m} min {s} s"
    if m: return f"{m} min {s} s"
    return f"{s} s"

def banner_inicio(fecha, hora, carpeta, total, changed, unchanged, errors):
    print(f"🚀 Procesando fecha {fecha} a las {hora}")
    print(_SEP)
    print(f"📂 Carpeta PDFs: {carpeta}")
    print(f"📊 Excel en:       {RUTA_EXCELS}")
    print(f"📌 Total estaciones: {total}")
    print(f"✅ Con cambios:        {changed}")
    print(f"🚫 Sin cambios:        {unchanged}")
    print(f"⚠️  Errores:         {errors}")
    print(_SEP)

# ─── FUNCIONES DE PROCESAMIENTO ─────────────────────────────────

def extraer_datos_ocr(ruta_pdf):
    client = vision.ImageAnnotatorClient()
    try:
        doc = fitz.open(ruta_pdf)
    except:
        return ""
    texto = ""
    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(2,2))
        img = vision.Image(content=pix.tobytes("png"))
        resp = client.document_text_detection(image=img)
        if resp.error.message: continue
        texto += resp.full_text_annotation.text + "\n"
    doc.close()
    return texto

def extraer_datos_pdf(ruta_pdf):
    texto = extraer_datos_ocr(ruta_pdf)
    folio_m = re.search(r"Folio\s*\n\s*([A-Z0-9]+)", texto, re.IGNORECASE)
    fecha_m = re.search(r"Fecha de firma.*?([0-9]{2}/[0-9]{2}/[0-9]{4})",
                        texto, re.IGNORECASE | re.DOTALL)
    return (
        folio_m.group(1).strip() if folio_m else None,
        fecha_m.group(1).strip() if fecha_m else None,
        []
    )

def obtener_carpetas_fechas(base, fecha_pdf):
    dates = []
    for d in os.listdir(base):
        if re.fullmatch(r"\d{8}", d):
            try:
                dt = datetime.strptime(d, "%d%m%Y").date()
                if dt == fecha_pdf:
                    dates.append(d)
            except: pass
    return sorted(dates)

def obtener_siguiente_fila_vacia(ws, start_row=2, cols=(1,2,3,4)):
    r = start_row
    while any(ws.cell(r,c).value not in (None,"") for c in cols):
        r += 1
    return r

def extract_pricing_from_email(fecha_obj=None):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox   = outlook.GetDefaultFolder(6)
    msgs    = inbox.Items
    sel_msg = None
    sel_time = None
    for m in msgs:
        try:
            if m.Class!=43: continue
            rd = m.ReceivedTime.date()
        except: continue
        if fecha_obj and rd!=fecha_obj: continue
        subj  = (m.Subject or "").lower()
        sender= m.SenderEmailAddress.lower()
        if "precios" in subj and (m.SenderName==SENDER_NAME or sender==SENDER_EMAIL):
            if sel_msg is None or m.ReceivedTime>sel_time:
                sel_msg, sel_time = m, m.ReceivedTime
    if not sel_msg: return []
    soup = BeautifulSoup(sel_msg.HTMLBody, 'html.parser')
    tbl  = soup.find('table')
    if not tbl: return []
    data=[]
    for tr in tbl.find_all('tr'):
        cols = tr.find_all(['td','th'])
        if len(cols)>=7 and cols[2].get_text(strip=True).startswith("BP "):
            data.append({
                "Estacion": cols[2].get_text(strip=True),
                "Regular":  _changed(cols[4]),
                "Premium":  _changed(cols[5]),
                "Diesel":   _changed(cols[6]),
            })
    return data

def _changed(cell):
    style = cell.get("style","").lower()
    txt   = cell.get_text(strip=True)
    if "background" in style and "white" not in style:
        return txt
    return None

def find_pricing_for_station(name, pricing_data):
    for rec in pricing_data:
        e = rec["Estacion"].lower()
        if name.lower() in e or e in name.lower():
            return rec
    return None

# ─── HILO DE PROCESAMIENTO ─────────────────────────────────────
class ProcesamientoThread(QThread):
    summary_signal = Signal(int,int,int,int)
    log_signal     = Signal(str)

    def __init__(self, fecha_str):
        super().__init__()
        self.fecha_str = fecha_str
        self.total = self.changed = self.unchanged = self.errors = 0

    def run(self):
        pythoncom.CoInitialize()
        try:
            try:
                fecha_busq = datetime.strptime(self.fecha_str, "%d/%m/%Y").date()
            except Exception as e:
                self.errors+=1
                self.log_signal.emit(f"Fecha inválida: {e}")
                self.summary_signal.emit(self.total,self.changed,self.unchanged,self.errors)
                return

            fecha_pdf = fecha_busq + timedelta(days=1)
            pricing = extract_pricing_from_email(fecha_busq)
            folders = obtener_carpetas_fechas(RUTA_BASE_PDFS, fecha_pdf)
            if not folders:
                self.log_signal.emit(f"No hay carpeta PDFs para {fecha_pdf:%d%m%Y}")
                return
            folder = folders[-1]
            self.log_signal.emit(f"Usando carpeta PDF: {folder}")

            existing_folios = set()
            # Leer folios ya en Excel
            for fn in os.listdir(RUTA_EXCELS):
                if fn.lower().endswith(".xlsx") and "copy" not in fn.lower():
                    wb = load_workbook(os.path.join(RUTA_EXCELS,fn))
                    for row in wb.active.iter_rows(min_row=2,values_only=True):
                        if row[2]:
                            existing_folios.add(str(row[2]).strip())
                    wb.close()

            for fn in os.listdir(RUTA_EXCELS):
                if not fn.lower().endswith(".xlsx") or "copy" in fn.lower():
                    continue
                self.total += 1
                base = fn.replace(" P.xlsx","").strip()
                self.log_signal.emit(_SEP)
                self.log_signal.emit(f"Procesando estación: {base}")
                wb = load_workbook(os.path.join(RUTA_EXCELS,fn))
                ws = wb.active

                # buscar PDFs
                pdfs = sorted(p for p in os.listdir(os.path.join(RUTA_BASE_PDFS,folder))
                              if p.lower().endswith(".pdf"))
                found_pdf=False; inserted=False
                for pdf in pdfs:
                    if base in pdf:
                        found_pdf=True
                        folio, date_firm, _ = extraer_datos_pdf(os.path.join(RUTA_BASE_PDFS,folder,pdf))
                        if folio and date_firm:
                            folio=folio.strip()
                            if folio in existing_folios:
                                self.log_signal.emit(f"Folio {folio} ya existe")
                            else:
                                self.log_signal.emit(f"Folio nuevo: {folio}")
                                dt = None
                                try:
                                    dt = datetime.strptime(date_firm, "%d/%m/%Y")
                                except:
                                    dt = date_firm
                                rec = find_pricing_for_station(base, pricing)
                                if rec:
                                    for prod in ("Regular","Premium","Diesel"):
                                        if rec.get(prod) is not None:
                                            nr = obtener_siguiente_fila_vacia(ws)
                                            ws.cell(nr,1).value = dt
                                            ws.cell(nr,2).value = rec[prod]
                                            ws.cell(nr,3).value = folio
                                            ws.cell(nr,4).value = prod
                                    existing_folios.add(folio)
                                    inserted=True
                                    self.changed+=1
                                else:
                                    self.log_signal.emit(f"No hay datos precios para {base}")
                        else:
                            self.log_signal.emit(f"No pudo extraer folio/fecha de {pdf}")
                if found_pdf and not inserted:
                    self.unchanged+=1

                # actualizar tabla y guardar
                try:
                    if "Table2" in ws.tables:
                        del ws.tables["Table2"]
                    tbl = Table(displayName="Table2", ref="A1:D30")
                    ts = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
                    tbl.tableStyleInfo = ts
                    ws.add_table(tbl)
                    wb.save(os.path.join(RUTA_EXCELS,fn))
                except Exception as e:
                    self.log_signal.emit(f"Error guardando {fn}: {e}")
                wb.close()

                self.summary_signal.emit(self.total,self.changed,self.unchanged,self.errors)

            self.log_signal.emit("✅ Proceso finalizado")
            banner_inicio(self.fecha_str, datetime.now().strftime("%H:%M"),
                          os.path.join(RUTA_BASE_PDFS,folder),
                          self.total,self.changed,self.unchanged,self.errors)

        finally:
            pythoncom.CoUninitialize()

# ─── INTERFAZ GRÁFICA ──────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestor de Folios – BP")
        self.resize(1200,700)
        w = QWidget(); self.setCentralWidget(w)
        grid = QGridLayout(w); grid.setContentsMargins(10,10,10,10)

        # Header
        hdr = QWidget(); hl = QHBoxLayout(hdr)
        lbl = QLabel("Fecha (dd/mm/yyyy):"); lbl.setFont(FUENTE_BASE)
        self.date_in = QLineEdit(); self.date_in.setFixedWidth(100)
        self.date_in.setText(date.today().strftime("%d/%m/%Y"))
        btn = QPushButton("Ejecutar"); btn.clicked.connect(self.start)
        btn.setStyleSheet(f"background:{COLOR_BOTON};color:white;")
        hl.addWidget(lbl); hl.addWidget(self.date_in); hl.addWidget(btn)
        grid.addWidget(hdr,0,0)

        # Divider
        div = QFrame(); div.setFrameShape(QFrame.VLine)
        div.setStyleSheet("background:#888;")
        grid.addWidget(div,0,1,2,1)

        # Log area
        self.log = QTextEdit(); self.log.setReadOnly(True)
        self.log.setFont(FUENTE_LOG)
        self.log.setStyleSheet(f"background:{COLOR_LOG_BG};color:{COLOR_LOG_FG}")
        grid.addWidget(self.log,1,0)

        # Summary panel
        panel = QWidget(); vl = QVBoxLayout(panel)
        self.lbl_tot = QLabel("📌 Analizadas: 0"); vl.addWidget(self.lbl_tot)
        self.lbl_chg = QLabel("✅ Cambios: 0"); vl.addWidget(self.lbl_chg)
        self.lbl_unc = QLabel("🚫 Sin cambios: 0"); vl.addWidget(self.lbl_unc)
        self.lbl_err = QLabel("⚠️ Errores: 0"); self.lbl_err.setStyleSheet(f"color:{COLOR_ERROR}")
        vl.addWidget(self.lbl_err)
        panel.setGraphicsEffect(QGraphicsDropShadowEffect(blurRadius=10))
        grid.addWidget(panel,0,2,2,1)

        # Signals & progress
        self.spinner = QProgressBar(); self.spinner.setRange(0,0); self.spinner.hide()
        grid.addWidget(self.spinner,2,0,1,3)

    @Slot(str)
    def append_log(self, txt): self.log.append(txt)

    @Slot(int,int,int,int)
    def update_summary(self, tot, ch, un, er):
        self.lbl_tot.setText(f"📌 Analizadas: {tot}")
        self.lbl_chg.setText(f"✅ Cambios: {ch}")
        self.lbl_unc.setText(f"🚫 Sin cambios: {un}")
        self.lbl_err.setText(f"⚠️ Errores: {er}")

    def start(self):
        self.spinner.show()
        thr = ProcesamientoThread(self.date_in.text())
        thr.log_signal.connect(self.append_log)
        thr.summary_signal.connect(self.update_summary)
        thr.finished.connect(lambda: self.spinner.hide())
        thr.start()

def main():
    app = QApplication(sys.argv)
    win = MainWindow(); win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
